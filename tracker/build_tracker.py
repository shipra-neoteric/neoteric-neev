import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

NAVY = "12355B"; ACCENT = "B3641C"; LIGHT = "EEF2F6"; ZEBRA = "F5F8FA"
BLUE_IN = "0000FF"; GREEN_LINK = "008000"

F = "Arial"
hdr = Font(name=F, size=10, bold=True, color="FFFFFF")
ttl = Font(name=F, size=14, bold=True, color=NAVY)
sub = Font(name=F, size=9, italic=True, color="5A5A5A")
bod = Font(name=F, size=10)
bodb = Font(name=F, size=10, bold=True)
inp = Font(name=F, size=10, color=BLUE_IN)
lnk = Font(name=F, size=10, color=GREEN_LINK)
fill_h = PatternFill("solid", fgColor=NAVY)
fill_z = PatternFill("solid", fgColor=ZEBRA)
fill_in = PatternFill("solid", fgColor="FFFFCC")
fill_l = PatternFill("solid", fgColor=LIGHT)
thin = Side(style="thin", color="B7C4CF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
ctr = Alignment(horizontal="center", vertical="center")
wrap = Alignment(wrap_text=True, vertical="top")

wb = openpyxl.Workbook()

TR = [f"T{i:02d}" for i in range(1, 13)]
N = len(TR)

DAYS = [
 ("D01","01-Sep-26","Tue"),("D02","02-Sep-26","Wed"),("D03","03-Sep-26","Thu"),("D04","04-Sep-26","Fri"),
 ("D05","05-Sep-26","Sat"),("D06","07-Sep-26","Mon"),("D07","08-Sep-26","Tue"),("D08","09-Sep-26","Wed"),
 ("D09","10-Sep-26","Thu"),("D10","11-Sep-26","Fri"),("D11","12-Sep-26","Sat"),("D12","14-Sep-26","Mon"),
 ("D13","15-Sep-26","Tue"),("D14","16-Sep-26","Wed"),("D15","17-Sep-26","Thu"),("D16","18-Sep-26","Fri"),
 ("D17","19-Sep-26","Sat"),("D18","21-Sep-26","Mon"),("D19","22-Sep-26","Tue"),("D20","23-Sep-26","Wed"),
 ("D21","24-Sep-26","Thu"),("D22","25-Sep-26","Fri"),("D23","26-Sep-26","Sat"),("D24","28-Sep-26","Mon"),
 ("D25","29-Sep-26","Tue"),("D26","30-Sep-26","Wed"),
]

def sheet(name, widths, title, subtitle=None):
    ws = wb.create_sheet(name)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws["A1"] = title; ws["A1"].font = ttl
    r = 2
    if subtitle:
        ws["A2"] = subtitle; ws["A2"].font = sub
        r = 3
    ws.freeze_panes = "A1"
    return ws, r + 1

def header_row(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=h)
        c.font = hdr; c.fill = fill_h; c.border = box; c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

def put(ws, row, col, val, font=bod, fill=None, align=None, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font; c.border = box
    if fill: c.fill = fill
    if align: c.alignment = align
    if fmt: c.number_format = fmt
    return c

# ============ 1. README ============
ws = wb.active; ws.title = "README"
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 100
ws["A1"] = "Project NEEV — Trainee Supervisor Programme Tracker"; ws["A1"].font = Font(name=F, size=16, bold=True, color=NAVY)
ws["A2"] = "Batch 2026-01 · 12 Trainee Supervisors · Phase 1: Sep–Dec 2026 · Gateway Assessment: Month 5 (Jan 2027) · Confirmation: 31 May 2027"; ws["A2"].font = sub
rows = [
 ("WHO DOES WHAT", ""),
 ("Rajat — runs it", "Training Coordinator. Fills Attendance and DailyLog at 17:30 every day, at the point of the log review — not at the end of the week. Keeps this workbook current."),
 ("Deepti — owns it", "Training Supervisor. Fills Weekly every Saturday, marks the D11 checkpoint and the D26 assessment, reads the bands, decides the department allotment, and takes it to the Saturday review."),
 ("Bharti — office", "Office Coordinator. Trainee names and joining documents into TraineeMaster, attendance data to payroll, training files."),
 ("", ""),
 ("WHEN TO FILL WHAT", ""),
 ("Daily 17:30", "Attendance and DailyLog. Two numbers per trainee. Five minutes."),
 ("Every Saturday", "Weekly — buddy rating out of 5 and module test out of 20."),
 ("12 Sep", "Checkpoint sheet. Velocity and Final compute themselves from it. Pods are formed from this result."),
 ("30 Sep", "Fill Capstone and department-block drill scores in Final. Read the band, allot the Month 2 immersion in the Immersions sheet."),
 ("", ""),
 ("CELL COLOURS", ""),
 ("Blue text on yellow", "You type here. These are the only cells you edit."),
 ("Black text", "A formula. Do not overwrite — if you type over it, the sheet stops working."),
 ("Green text", "A link pulling a value from another sheet."),
 ("", ""),
 ("THE TWO THINGS WE MEASURE", ""),
 ("Score", "What the trainee knows and can do. The weighted total in the Final sheet, out of 100."),
 ("Learning Velocity", "How fast they are learning, measured as normalised gain: (Gate − Baseline) ÷ (100 − Baseline) × 100."),
 ("Why normalised", "A trainee who moves 30→70 captured 57% of the improvement available to them. One who moves 65→75 captured 29%. The second has a higher score and the lower velocity. Velocity, not starting position, drives the Month 2 assignment."),
 ("", ""),
 ("BANDS", ""),
 ("A — Fast", "Final >= 75 and Velocity >= 55. Given the harder scope in each immersion; watch as a future department lead."),
 ("B — On track", "Final >= 60. Standard immersion scope with daily supervision. Most people should be here."),
 ("C — Developing", "Final >= 45. Written improvement plan naming two specific gaps, a named mentor, re-check next month."),
 ("D — At risk", "Final < 45. Honest written conversation with Deepti. NOT an exit — the Gateway in Month 5 is the exit point. But they must be told in September, not discover it in January."),
 ("", ""),
 ("ASSUMPTIONS", ""),
 ("Hours", "09:00 to 18:00 per the offer letter. Eight scheduled early-muster days (2, 5, 8, 14, 17, 21, 24, 28 Sep) start 07:30 and end 16:30."),
 ("Offs", "4 offs per month per the offer letter. In September these are the four Sundays, giving 26 working days."),
 ("Weights", "Set by Rahul and Deepti; stated in Part A6 of the Trainee Handbook. Editable in the Final sheet, row 4."),
 ("Baseline", "The D01 diagnostic is NOT used for ranking. It exists only so that improvement can be measured. Tell the trainees this."),
 ("Gateway", "Month 5, January 2027. Theory + practical. This workbook covers Month 1; the same structure extends through Months 2-4."),
]
r = 4
for a, b in rows:
    ca = ws.cell(row=r, column=1, value=a)
    cb = ws.cell(row=r, column=2, value=b)
    if b == "":
        ca.font = Font(name=F, size=11, bold=True, color=ACCENT)
    else:
        ca.font = bodb; cb.font = bod; cb.alignment = wrap
    r += 1
ws.row_dimensions[19].height = 45
ws.row_dimensions[27].height = 28

# ============ 2. TraineeMaster ============
ws, r = sheet("TraineeMaster", [8, 26, 16, 10, 24, 14, 26],
              "Trainee Master",
              "Bharti fills columns B, C and G at joining. Pod and Site Buddy are assigned on 12 Sep after the checkpoint (see Rotation sheet). Four pods of three, one buddy per pod.")
header_row(ws, r, ["ID", "Name", "Diploma / branch", "Pod", "Site Buddy", "Base site", "Notes"])
hr = r
for i, t in enumerate(TR):
    rr = hr + 1 + i
    put(ws, rr, 1, t, bodb, align=ctr)
    for col in (2, 3, 4, 5, 7):
        put(ws, rr, col, "Rahul Sharma (example)" if (col == 2 and i == 0) else None, inp, fill_in)
    put(ws, rr, 6, "Zen Garden", bod)
    if i == 0:
        ws.cell(row=rr, column=3, value="Civil — Govt Polytechnic Gwalior")
        ws.cell(row=rr, column=4, value="A")
        ws.cell(row=rr, column=5, value="Sr. Engineer — Structures")
ws.cell(row=hr + 1 + N + 1, column=1, value="Example row shown in T01. Overwrite it with the real trainee.").font = sub

# ============ 3. Calendar ============
ws, r = sheet("Calendar", [8, 13, 8, 12, 60, 26],
              "Programme Calendar — Month 1",
              "26 working days, 1–30 September 2026. Sundays off. Adjust for any local holiday before Day 1.")
header_row(ws, r, ["Day", "Date", "Weekday", "Phase", "Focus / classroom module (15:30–17:00)", "Owner"])
hr = r
focus = [
 "Induction, safety, code of conduct · BASELINE DIAGNOSTIC (D01)", "HR + Safety Officer + TC",
 "Reading drawings I — architectural set, grids, levels, symbols", "Design / Planning Head",
 "Reading drawings II — structural set, column & beam schedules, BBS", "Structural / PMC Lead",
 "Measurement, mensuration, IS 1200 and the Measurement Book", "Sr. QS",
 "Land survey & setting out — total station, grids, profiles", "Site Surveyor",
 "Levels & levelling — auto level, RL, level book, the 1-metre line", "Site Surveyor",
 "Cement — types, field tests, lab tests, storage, consumption control", "QC Head",
 "Concrete & reinforcement — mix, slump, cubes, BBS, cover", "QC Head + Sr. Engineer",
 "Shuttering, masonry & block machinery; stores walkthrough", "Sr. Engineer — Structures",
 "Estimation, BOQ, rate analysis; procurement cycle", "Sr. QS + Purchase Head",
 "QC system, NCR, SWA · MONTH 1 CHECKPOINT · pods formed (D11)", "QC Head + Deepti",
 "Dept taster B1 begins — P1 Supervision, P2 Quality, P3 Measurement, P4 Store", "Rajat + dept heads",
 "Department taster block", "Dept head", "Department taster block", "Dept head",
 "Department taster block", "Dept head", "Department taster block", "Dept head",
 "Taster B1 ends · Saturday review with Deepti and Sponsor", "Deepti",
 "Taster B2 begins — P1 Quality, P2 Measurement, P3 Store, P4 Supervision", "Rajat + dept heads",
 "Department taster block", "Dept head", "Department taster block", "Dept head",
 "Department taster block", "Dept head",
 "Taster B2 ends · Saturday review with Deepti and Sponsor", "Deepti",
 "Taster B3 begins — P1 Measurement, P2 Store, P3 Supervision, P4 Quality", "Rajat + dept heads",
 "Department taster block", "Dept head", "Department taster block", "Dept head",
 "Taster B4 · capstone preparation", "Rajat",
 "MONTH 1 ASSESSMENT · capstone · department aptitude read · Month 2 immersion allotted", "Rahul + Deepti + dept heads",
]
for i, (dcode, dt, wd) in enumerate(DAYS):
    rr = hr + 1 + i
    phase = "Phase 1" if i < 11 else "Phase 2"
    f = fill_z if i % 2 else None
    put(ws, rr, 1, dcode, bodb, f, ctr)
    put(ws, rr, 2, dt, bod, f, ctr)
    put(ws, rr, 3, wd, bod, f, ctr)
    put(ws, rr, 4, phase, bod, f, ctr)
    put(ws, rr, 5, focus[i * 2], bod, f)
    put(ws, rr, 6, focus[i * 2 + 1], bod, f)

r2 = hr + 1 + len(DAYS) + 2
ws.cell(row=r2, column=1, value="Month 2 — Supervised Ownership (1–31 October 2026)").font = Font(name=F, size=12, bold=True, color=NAVY)
header_row(ws, r2 + 1, ["Day", "Date", "Weekday", "Phase", "Milestone", "Owner"])
m2 = [
 ("D27", "01-Oct-26", "Thu", "Month 2", "Scope handover — each trainee receives a named, bounded scope and a reporting engineer", "Programme Sponsor"),
 ("—", "03-Oct-26", "Sat", "Month 2", "Week 1 review — first independent DPR and checklist submitted", "TC"),
 ("—", "10-Oct-26", "Sat", "Month 2", "Week 2 review — first independent measurement and NCR", "TC"),
 ("—", "17-Oct-26", "Sat", "Month 2", "Week 3 review — contractor interaction and productivity tracking", "TC"),
 ("D45", "22-Oct-26", "Thu", "Month 2", "MID-POINT REVIEW — band re-confirmed or revised, improvement plans reset", "Sponsor + TC"),
 ("—", "24-Oct-26", "Sat", "Month 2", "Week 4 review — snagging and handover exposure", "TC"),
 ("D60", "31-Oct-26", "Sat", "Month 2", "CONFIRMATION REVIEW — 60-day checklist signed, certification, role placement, letters issued", "CEO + Sponsor + HR"),
]
for i, row in enumerate(m2):
    rr = r2 + 2 + i
    f = fill_z if i % 2 else None
    for j, v in enumerate(row, 1):
        put(ws, rr, j, v, bodb if j == 1 else bod, f, ctr if j in (1, 2, 3, 4) else None)

# ============ 4. Rotation ============
ws, r = sheet("Rotation", [13, 22, 19, 19, 19, 19],
              "Pods, Departments & Rotation",
              "Four pods of three. Four departments. Read a row across: in block B1, Pod 1 is in Supervision while Pod 2 is in Quality, Pod 3 in Measurement and Pod 4 in Store. Then everyone moves one place along.")
header_row(ws, r, ["Block", "Dates", "Pod 1", "Pod 2", "Pod 3", "Pod 4"])
hr = r
taster = [
 ("B1", "Mon 14 - Thu 17 Sep", "Supervision", "Quality", "Measurement", "Store"),
 ("B2", "Fri 18 - Tue 22 Sep", "Quality", "Measurement", "Store", "Supervision"),
 ("B3", "Wed 23 - Fri 25 Sep", "Measurement", "Store", "Supervision", "Quality"),
 ("B4", "Sat 26 - Tue 29 Sep", "Store", "Supervision", "Quality", "Measurement"),
 ("D26", "Wed 30 Sep", "Month 1 assessment", "capstone", "aptitude read", "Month 2 allotted"),
]
for i2, row in enumerate(taster):
    rr = hr + 1 + i2
    f = fill_z if i2 % 2 else None
    for j2, v in enumerate(row, 1):
        put(ws, rr, j2, v, bodb if j2 <= 2 else bod, f, ctr if j2 != 2 else None)

r2 = hr + 7
ws.cell(row=r2, column=1, value="Months 2 to 4 - department immersions (same pods, a whole month each)").font = Font(name=F, size=12, bold=True, color=NAVY)
header_row(ws, r2 + 1, ["Period", "When", "Pod 1", "Pod 2", "Pod 3", "Pod 4"])
imm = [
 ("Month 2", "October 2026", "Supervision", "Quality", "Measurement", "Store"),
 ("Month 3", "November 2026", "Quality", "Measurement", "Store", "Supervision"),
 ("Month 4", "December 2026", "Measurement", "Store", "Supervision", "Quality"),
 ("Last week", "December 2026", "Store", "Supervision", "Quality", "Measurement"),
]
for i2, row in enumerate(imm):
    rr = r2 + 2 + i2
    f = fill_z if i2 % 2 else None
    for j2, v in enumerate(row, 1):
        put(ws, rr, j2, v, bodb if j2 <= 2 else bod, f, ctr if j2 != 2 else None)
ws.cell(row=r2 + 6, column=1, value="Three months deep in three departments, then one week in the fourth so nobody reaches the Gateway with a blank space. Every department hosts exactly one pod at a time, which is what makes this affordable in supervision time.").font = sub

r3 = r2 + 8
ws.cell(row=r3, column=1, value="The four departments").font = Font(name=F, size=12, bold=True, color=NAVY)
header_row(ws, r3 + 1, ["Dept", "What the job is", "Handbook modules", "Department head", "", ""])
depts = [
 ("Supervision", "Getting work built correctly and on time, by people who do not report to you.", "M04 M05 M08 M09 M10 M11 M13 M14 M15 M21", "Sr. Engineer (Structures) + PM"),
 ("Quality", "Deciding whether what was built is acceptable, and stopping it when it is not.", "M01 M06 M07 M12 M19 M22", "QC Head"),
 ("Measurement", "Turning built work into numbers that can be paid.", "M02 M03 M17 M18", "Senior QS"),
 ("Store", "Controlling the 55-70% of project cost that passes through the store.", "M16 + procurement cycle", "Store In-charge + Purchase Head"),
]
ws.column_dimensions["B"].width = 52
for i2, row in enumerate(depts):
    rr = r3 + 2 + i2
    f = fill_z if i2 % 2 else None
    for j2, v in enumerate(row, 1):
        c = put(ws, rr, j2, v, bodb if j2 == 1 else bod, f)
        if j2 == 2: c.alignment = wrap
    ws.row_dimensions[rr].height = 28

r4 = r3 + 7
ws.cell(row=r4, column=1, value="Pod assignment - fill on 12 Sep, after the checkpoint").font = Font(name=F, size=12, bold=True, color=NAVY)
header_row(ws, r4 + 1, ["Pod", "Trainee 1", "Trainee 2", "Trainee 3", "Site Buddy", "Notes"])
for i2 in range(4):
    rr = r4 + 2 + i2
    put(ws, rr, 1, f"Pod {i2+1}", bodb, align=ctr)
    for j2 in range(2, 7):
        put(ws, rr, j2, None, inp, fill_in)
nr = r4 + 7
for i2, line in enumerate([
 "POD COMPOSITION RULE",
 "One strong, one middle and one developing trainee in every pod. Do not put the three quickest together - you get one pod that races ahead, three that stall, and no peer teaching anywhere.",
 "One site buddy per pod of three. Three is the maximum one person can genuinely supervise on a live site; it is a safety limit as much as a teaching one.",
 "Pods stay together for the whole of Phase 1. Rebalance only if a pod is clearly dragging, and only at a month boundary.",
]):
    c = ws.cell(row=nr + i2, column=1, value=line)
    c.font = Font(name=F, size=10, bold=(i2 == 0), color=ACCENT if i2 == 0 else "222222")

# ============ 5. TrainerMatrix ============
ws, r = sheet("TrainerMatrix", [8, 40, 30, 26, 14, 40],
              "Trainer & Ownership Matrix",
              "Module numbers match the Trainee Handbook (M01–M22). Recommended by ROLE — fill column D with the actual person. Read the selection rule at the bottom before you assign anyone.")
header_row(ws, r, ["Ref", "Module", "Recommended trainer (role)", "Assigned name", "Hours", "Why this role"])
hr = r
tm = [
 ("M01","Site Safety & EHS","Safety Officer / EHS Lead","",  "3", "Owns the standard and the permits; must be the face of it from Day 1"),
 ("M02","Reading Drawings (D02 architectural + D03 structural & BBS)","Design/Planning Head + Structural Engineer","", "6", "Two days, two teachers: the Planning Head issues the drawings, the structural engineer explains consultant intent"),
 ("M03","Measurement, IS 1200 & the Measurement Book","Senior Quantity Surveyor","", "3", "MB discipline is a QS competency, not a site-engineer one"),
 ("M04","Land Survey & Setting Out","Site Surveyor","", "3", "Instrument skill can only be taught hands-on by the person who uses it daily"),
 ("M05","Levels & Levelling","Site Surveyor","", "3", "Same instrument, same person, continuity of method"),
 ("M06","Cement & Cement Testing","QC Head / Lab In-charge","", "3", "Owns the test regime and the acceptance decision"),
 ("M07","Concrete & Concrete Testing","QC Head + Sr. Engineer (Structures)","", "3", "Theory from QC, pour discipline from the engineer who runs pours"),
 ("M08","Reinforcement Steel & BBS","Sr. Engineer — Structures","", "3", "Bar fixing supervision is a daily execution skill"),
 ("M09","Shuttering & Formwork","Sr. Engineer — Structures / Formwork In-charge","", "3", "Deshuttering judgement cannot be taught from a table"),
 ("M10","Brick, Block & Masonry Machinery","Sr. Engineer — Finishing","", "3", "Owns the trade and the machinery on site"),
 ("M11","Plastering","Sr. Engineer — Finishing","", "2", "Dots, screeds and tap-testing are learned by hand"),
 ("M12","Waterproofing","QC Head + waterproofing agency","", "3", "The agency knows the system; QC owns the acceptance and the record"),
 ("M13","Tiling & Flooring","Sr. Engineer — Finishing","", "2", "Layout judgement is the whole skill"),
 ("M14","Plumbing & Sanitary","MEP Engineer","", "3", "Test discipline before concealment is the entire module"),
 ("M15","Electrical","MEP Engineer / Licensed Electrical Engineer","", "3", "Statutory competence required; do not delegate below this level"),
 ("M16","Stores & Inventory Management","Store In-charge + Purchase Head","", "4", "Physical process from the store, commercial process from purchase"),
 ("M17","Estimation, BOQ & Rate Analysis","Senior Quantity Surveyor","", "4", "Rate analysis is where site engineers are weakest and most exploited"),
 ("M18","Quantity Surveying & Contractor Billing","Senior QS + Accounts","", "3", "Deductions and recoveries need the accounts view too"),
 ("M19","QC — Checklists, NCR & Stop-Work Authority","QC Head + Programme Sponsor","", "3", "Policy weight requires the Sponsor in the room, not just the QC Head"),
 ("M20","Site Documentation, Reporting & Digital Discipline","Planning Engineer + TC","", "3", "TC must own this because TC enforces it every evening"),
 ("M21","Labour, Contractor & Site Management","Project Manager","", "3", "Only someone who has actually managed contractors can teach this credibly"),
 ("M22","Snagging, Handover & Customer Quality","Quality Head + CRM/Handover Lead","", "3", "The customer-facing view is what makes it land"),
]
for i, row in enumerate(tm):
    rr = hr + 1 + i
    f = fill_z if i % 2 else None
    for j, v in enumerate(row, 1):
        if j == 4:
            put(ws, rr, j, None, inp, fill_in)
        else:
            c = put(ws, rr, j, int(v) if j == 5 else v, bodb if j == 1 else bod, f, ctr if j in (1, 5) else None)
            if j == 6: c.alignment = wrap
tot = hr + 1 + len(tm)
put(ws, tot, 2, "Total trainer hours across the programme", bodb, fill_l)
put(ws, tot, 5, f"=SUM(E{hr+1}:E{tot-1})", bodb, fill_l, ctr)
put(ws, tot, 6, "Roughly 5–7 working days of senior time in total, spread across ~12 people. That is the real cost of this programme.", sub, fill_l).alignment = wrap

note_r = tot + 2
for i, line in enumerate([
 "TRAINER SELECTION RULE — read before assigning anyone",
 "1. Do not assign anyone who has ever been the subject of a show-cause notice on checklist or inspection discipline. A trainer transmits habits, not just content.",
 "2. Do not assign a trainer to a module they have not personally executed in the last twelve months.",
 "3. Each trainer gets a fixed 90-minute afternoon window and one field drill. They are not pulled off site for a day. This is what makes the programme affordable.",
 "4. The Training Coordinator sits in on every session in the first week and spot-checks after that. Trainer quality is audited, not assumed.",
 "5. If a role is vacant or the person is unavailable, the Programme Sponsor teaches it. The module does not get dropped and it does not get moved.",
]):
    c = ws.cell(row=note_r + i, column=1, value=line)
    c.font = Font(name=F, size=10, bold=(i == 0), color=ACCENT if i == 0 else "222222")

# ============ 6. Attendance ============
ws, r = sheet("Attendance", [8, 13, 9] + [6] * N + [10],
              "Attendance Register",
              "Enter P (present), A (absent), L (leave), H (half day). Filled at 17:30 daily by the Training Coordinator.")
header_row(ws, r, ["Day", "Date", "Wd"] + TR + ["% Present"])
hr = r
for i, (dc, dt, wd) in enumerate(DAYS):
    rr = hr + 1 + i
    f = fill_z if i % 2 else None
    put(ws, rr, 1, dc, bodb, f, ctr); put(ws, rr, 2, dt, bod, f, ctr); put(ws, rr, 3, wd, bod, f, ctr)
    for j in range(N):
        put(ws, rr, 4 + j, "P" if i == 0 else None, inp, fill_in, ctr)
    put(ws, rr, 4 + N, f'=IFERROR(COUNTIF(D{rr}:{get_column_letter(3+N)}{rr},"P")/{N},"")', bod, f, ctr, "0%")
sr = hr + 1 + len(DAYS)
put(ws, sr, 3, "Days present", bodb, fill_l, ctr)
for j in range(N):
    cl = get_column_letter(4 + j)
    put(ws, sr, 4 + j, f'=COUNTIF({cl}{hr+1}:{cl}{sr-1},"P")+0.5*COUNTIF({cl}{hr+1}:{cl}{sr-1},"H")', bodb, fill_l, ctr)
put(ws, sr + 1, 3, "Attendance %", bodb, fill_l, ctr)
for j in range(N):
    cl = get_column_letter(4 + j)
    put(ws, sr + 1, 4 + j, f"={cl}{sr}/{len(DAYS)}", bodb, fill_l, ctr, "0%")

# ============ 7. DailyLog ============
ws, r = sheet("DailyLog", [8, 13, 9] + [6] * N + [10],
              "Daily Log Quality Score",
              "Rated 1–5 by the Training Coordinator at the 17:00 log review, in front of the trainee. 5 = a colleague who was not on site could reconstruct the day from it. 1 = no numbers, no locations, no question.")
header_row(ws, r, ["Day", "Date", "Wd"] + TR + ["Batch avg"])
hr = r
for i, (dc, dt, wd) in enumerate(DAYS):
    rr = hr + 1 + i
    f = fill_z if i % 2 else None
    put(ws, rr, 1, dc, bodb, f, ctr); put(ws, rr, 2, dt, bod, f, ctr); put(ws, rr, 3, wd, bod, f, ctr)
    for j in range(N):
        put(ws, rr, 4 + j, 3 if i == 0 else None, inp, fill_in, ctr)
    put(ws, rr, 4 + N, f"=IFERROR(AVERAGE(D{rr}:{get_column_letter(3+N)}{rr}),\"\")", bod, f, ctr, "0.0")
sr = hr + 1 + len(DAYS)
put(ws, sr, 3, "Average", bodb, fill_l, ctr)
for j in range(N):
    cl = get_column_letter(4 + j)
    put(ws, sr, 4 + j, f'=IFERROR(AVERAGE({cl}{hr+1}:{cl}{sr-1}),"")', bodb, fill_l, ctr, "0.00")
put(ws, sr + 1, 3, "Trend (last 8 vs first 8)", bodb, fill_l, ctr)
for j in range(N):
    cl = get_column_letter(4 + j)
    put(ws, sr + 1, 4 + j, f'=IFERROR(AVERAGE({cl}{sr-8}:{cl}{sr-1})-AVERAGE({cl}{hr+1}:{cl}{hr+8}),"")', bodb, fill_l, ctr, "+0.00;-0.00;0.00")
ws.cell(row=sr + 3, column=1, value="A positive trend is the point. A trainee whose log score is flat for three weeks is not being corrected — that is a Training Coordinator failure, not a trainee failure.").font = sub
dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"D{hr+1}:{get_column_letter(3+N)}{sr-1}")

# ============ 8. Weekly ============
ws, r = sheet("Weekly", [8, 26] + [11] * 8 + [12, 12],
              "Weekly Ratings",
              "Site Buddy rating (1–5: initiative, discipline, safety behaviour, punctuality, how they take correction) and the weekly module test (out of 20). Filled every Saturday before the Sponsor review.")
weeks = ["Wk1 (1–5 Sep)", "Wk2 (7–12 Sep)", "Wk3 (14–19 Sep)", "Wk4 (21–30 Sep)"]
header_row(ws, r, ["ID", "Name"] + [f"{w}\n{k}" for w in ["Wk1", "Wk2", "Wk3", "Wk4"] for k in ["Buddy /5", "Test /20"]] + ["Buddy avg", "Test avg %"])
hr = r
for i, t in enumerate(TR):
    rr = hr + 1 + i
    put(ws, rr, 1, t, bodb, align=ctr)
    put(ws, rr, 2, f"=IFERROR(INDEX(TraineeMaster!$B${hr}:$B${hr+N},MATCH($A{rr},TraineeMaster!$A${hr}:$A${hr+N},0)),\"\")", lnk)
    for j in range(8):
        put(ws, rr, 3 + j, (4 if j % 2 == 0 else 14) if i == 0 else None, inp, fill_in, ctr)
    put(ws, rr, 11, f'=IFERROR(AVERAGE(C{rr},E{rr},G{rr},I{rr}),"")', bod, align=ctr, fmt="0.00")
    put(ws, rr, 12, f'=IFERROR(AVERAGE(D{rr},F{rr},H{rr},J{rr})/20,"")', bod, align=ctr, fmt="0%")
ws.cell(row=hr + N + 2, column=2, value="Note: Wk1 has 5 working days, Wk2 has 6, Wk3 has 6, Wk4 has 9. The averages are per-week, not per-day, so the weeks weigh equally.").font = sub

# ============ 9. GateD11 ============
ws, r = sheet("GateD11", [8, 26, 14, 14, 14, 14, 12, 14, 46],
              "Month 1 Checkpoint Assessment — D11, Saturday 12 September 2026",
              "100 marks: 40 written (from the handbook self-check questions), 30 practical field task, 30 behavioural & situational. Marked by the QC Head and Deepti independently, then reconciled. This is a checkpoint, not a gate — the Gateway is Month 5. Pods are formed from this result.")
header_row(ws, r, ["ID", "Name", "Baseline D01 /100", "Written /40", "Practical /30", "Behavioural /30", "Checkpt /100", "Velocity", "Marker's one-line judgement"])
hr = r
for i, t in enumerate(TR):
    rr = hr + 1 + i
    f = fill_z if i % 2 else None
    put(ws, rr, 1, t, bodb, f, ctr)
    put(ws, rr, 2, f"=IFERROR(INDEX(TraineeMaster!$B$5:$B${4+N},MATCH($A{rr},TraineeMaster!$A$5:$A${4+N},0)),\"\")", lnk, f)
    for j in (3, 4, 5, 6):
        put(ws, rr, j, [32, 27, 21, 22][j - 3] if i == 0 else None, inp, fill_in, ctr)
    put(ws, rr, 7, f'=IFERROR(D{rr}+E{rr}+F{rr},"")', bodb, f, ctr, "0")
    put(ws, rr, 8, f'=IFERROR(IF(C{rr}>=100,"",(G{rr}-C{rr})/(100-C{rr})*100),"")', bodb, f, ctr, "0")
    put(ws, rr, 9, None, inp, fill_in).alignment = wrap
    ws.row_dimensions[rr].height = 26
ws.cell(row=hr + N + 2, column=2, value="Velocity = normalised gain = (Gate − Baseline) ÷ (100 − Baseline) × 100. It measures the share of available improvement the trainee actually captured, which is what 'learning speed' means once you control for where they started.").font = sub
ws.cell(row=hr + N + 3, column=2, value="The one-line judgement column is mandatory. A number without a sentence tells you nothing you can act on in the D11 conversation.").font = sub

# ============ 10. Final ============
ws, r = sheet("Final", [8, 24, 11, 11, 11, 11, 11, 11, 11, 11, 10, 34],
              "Month 1 Final Scorecard — D26, Wednesday 30 September 2026",
              "Weighted total per the weights in row 4 (editable, blue). Bands and Month 2 scope levels follow the rule in the README. Feeds the department aptitude read and the Month 2 immersion allotment.")
ws["A4"] = "Weights →"; ws["A4"].font = bodb
wcells = {"C": 0.15, "D": 0.15, "E": 0.15, "F": 0.25, "G": 0.15, "H": 0.15}
for col, v in wcells.items():
    c = ws[f"{col}4"]; c.value = v; c.font = inp; c.fill = fill_in; c.border = box; c.alignment = ctr; c.number_format = "0%"
ws["I4"] = "=SUM(C4:H4)"; ws["I4"].font = bodb; ws["I4"].alignment = ctr; ws["I4"].number_format = "0%"; ws["I4"].border = box
ws["J4"] = "must total 100%"; ws["J4"].font = sub
hr = 6
header_row(ws, hr, ["ID", "Name", "Log /5", "Buddy /5", "Test /20", "Checkpt /100", "Dept drills /100", "Capstone /100", "FINAL /100", "Velocity", "Band", "Month 2 scope level"])
for i, t in enumerate(TR):
    rr = hr + 1 + i
    f = fill_z if i % 2 else None
    put(ws, rr, 1, t, bodb, f, ctr)
    put(ws, rr, 2, f"=IFERROR(INDEX(TraineeMaster!$B$5:$B${4+N},MATCH($A{rr},TraineeMaster!$A$5:$A${4+N},0)),\"\")", lnk, f)
    dcol = get_column_letter(4 + i)
    put(ws, rr, 3, f'=IFERROR(DailyLog!{dcol}${4+1+len(DAYS)},"")', lnk, f, ctr, "0.00")
    put(ws, rr, 4, f'=IFERROR(Weekly!$K{5+i},"")', lnk, f, ctr, "0.00")
    put(ws, rr, 5, f'=IFERROR(Weekly!$L{5+i}*20,"")', lnk, f, ctr, "0.0")
    put(ws, rr, 6, f'=IFERROR(GateD11!$G{5+i},"")', lnk, f, ctr, "0")
    for j in (7, 8):
        put(ws, rr, j, [72, 68][j - 7] if i == 0 else None, inp, fill_in, ctr)
    put(ws, rr, 9, (f'=IFERROR(C{rr}/5*100*$C$4 + D{rr}/5*100*$D$4 + E{rr}/20*100*$E$4 '
                    f'+ F{rr}*$F$4 + G{rr}*$G$4 + H{rr}*$H$4,"")'), bodb, f, ctr, "0.0")
    put(ws, rr, 10, f'=IFERROR(GateD11!$H{5+i},"")', lnk, f, ctr, "0")
    put(ws, rr, 11, f'=IFERROR(IF(AND(I{rr}>=75,J{rr}>=55),"A",IF(I{rr}>=60,"B",IF(I{rr}>=45,"C","D"))),"")', bodb, f, ctr)
    put(ws, rr, 12, f'=IFERROR(IF(K{rr}="A","Harder immersion scope, weekly supervision",'
                    f'IF(K{rr}="B","Standard immersion scope, daily supervision",'
                    f'IF(K{rr}="C","Improvement plan naming two gaps + named mentor, re-check next month",'
                    f'IF(K{rr}="D","Written conversation with Deepti now - NOT an exit, Gateway is the exit point","")))),"")', bod, f)
    ws.row_dimensions[rr].height = 24
last = hr + N
ws.conditional_formatting.add(f"K{hr+1}:K{last}",
    CellIsRule(operator="equal", formula=['"A"'], fill=PatternFill("solid", fgColor="D6EAD6")))
ws.conditional_formatting.add(f"K{hr+1}:K{last}",
    CellIsRule(operator="equal", formula=['"C"'], fill=PatternFill("solid", fgColor="FDF0D5")))
ws.conditional_formatting.add(f"K{hr+1}:K{last}",
    CellIsRule(operator="equal", formula=['"D"'], fill=PatternFill("solid", fgColor="F6D9D9")))
nr = last + 2
for i, line in enumerate([
 "Dept drills /100 = the average of the four field-drill outputs from the September department blocks, marked by each department head.",
 "Capstone /100 = the 10-minute D26 presentation on one real problem found on site and the recommendation, marked by Rahul, the Sponsor and Deepti.",
 "A trainee in band D at D26 should already have heard it at D11. The Gateway in Month 5 is the exit point - but nobody should reach it surprised.",
]):
    ws.cell(row=nr + i, column=2, value=line).font = sub

# ============ 11. Immersions ============
ws, r = sheet("Immersions", [8, 22, 15, 15, 15, 24, 40],
              "Months 2 to 4 - Department Immersions",
              "Filled at the 30 Sep placement meeting. The department columns follow the pod rotation in the Rotation sheet; the scope and reporting engineer are decided per trainee.")
header_row(ws, r, ["ID", "Name", "Month 2 dept", "Month 3 dept", "Month 4 dept", "Reporting engineer", "Month 2 scope (be specific)"])
hr = r
for i2, t in enumerate(TR):
    rr = hr + 1 + i2
    f = fill_z if i2 % 2 else None
    put(ws, rr, 1, t, bodb, f, ctr)
    put(ws, rr, 2, f"=IFERROR(INDEX(TraineeMaster!$B$5:$B${4+N},MATCH($A{rr},TraineeMaster!$A$5:$A${4+N},0)),\"\")", lnk, f)
    for j2 in (3, 4, 5, 6):
        put(ws, rr, j2, None, inp, fill_in, ctr if j2 <= 5 else None)
    put(ws, rr, 7, "Zen Garden Tower 1, floors 4-5, masonry and plaster (example)" if i2 == 0 else None, inp, fill_in)
    ws.row_dimensions[rr].height = 24

nr = hr + N + 2
for i2, line in enumerate([
 "SCOPE DESIGN RULE",
 "Small enough that the trainee can actually control it; real enough that getting it wrong shows. One floor of one trade is right. 'Assist the Site Engineer' is not a scope and teaches nothing - it is the default every trainee programme drifts into, and it is why most of them produce nothing.",
 "What ownership looks like in each department: in Store, they own the reconciliation of two materials for the month and sign it. In Quality, the checklists for one floor. In Measurement, the take-off for one flat type. In Supervision, one trade on one floor.",
 "Write the reporting engineer's accountability for their trainee into that engineer's own monthly objectives. Otherwise it competes with site work and loses.",
 "Each immersion ends with a written review by the department head. Three independent opinions per trainee by December is what makes the Gateway allotment defensible.",
]):
    c = ws.cell(row=nr + i2, column=2, value=line)
    c.font = Font(name=F, size=10, bold=(i2 == 0), color=ACCENT if i2 == 0 else "222222")
    c.alignment = wrap

r2 = nr + 7
ws.cell(row=r2, column=2, value="Phase 1 milestones after September").font = Font(name=F, size=12, bold=True, color=NAVY)
header_row(ws, r2 + 1, ["", "When", "Milestone", "Owner", "", "", ""])
ms = [
 ("31 Oct 2026", "Month 2 immersion review - written, by the department head", "Dept head + Deepti"),
 ("30 Nov 2026", "Month 3 immersion review", "Dept head + Deepti"),
 ("24 Dec 2026", "Month 4 immersion review; catch-up week in the fourth department begins", "Dept head + Deepti"),
 ("31 Dec 2026", "Gateway briefing - every trainee told exactly what the assessment covers", "Deepti"),
 ("Week 1 Jan 2027", "GATEWAY ASSESSMENT - theory and practical", "Rahul + Deepti + QC Head"),
 ("Jan 2027", "Department allotment issued in writing", "Rahul + HR"),
 ("31 May 2027", "End of 9-month probation - confirmation letters and revised terms", "Rahul + HR"),
]
for i2, row in enumerate(ms):
    rr = r2 + 2 + i2
    f = fill_z if i2 % 2 else None
    put(ws, rr, 2, row[0], bodb, f)
    put(ws, rr, 3, row[1], bod, f)
    put(ws, rr, 4, row[2], bod, f)

wb.save("../dist/NEEV_Programme_Tracker.xlsx")
print("saved")
